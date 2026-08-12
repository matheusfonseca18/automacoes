import pandas as pd
from pandas import Timestamp
from openpyxl import load_workbook
from openpyxl.utils import range_boundaries, get_column_letter
import win32com.client as win32
from dotenv import load_dotenv
import os
from pathlib import Path
import time
import shutil
from datetime import datetime
from shared.backup_utils import fazer_backup

def executar(logger):

    inicio = time.perf_counter()

    # Carregar variáveis de ambiente
    load_dotenv()
    destinatario = os.getenv("destinatario")
    cc = os.getenv("cc")
    #Caminhos
    base = Path(os.environ["OneDrive"]) # acha o caminho variável do OneDrive
    planilha_baixada_path = base / os.getenv("baixada_montar")
    planilha_final_path = os.getenv("planilha_final_path") # caminho final na rede

    # funções
    def obter_estrutura_tabela(ws):
        tabela = ws.tables[list(ws.tables.keys())[0]]
        min_col, min_row, max_col, max_row = range_boundaries(tabela.ref)
        headers = [ws.cell(row=min_row, column=c).value for c in range(min_col, max_col + 1)]
        
        colunas_data = {}
        for j, col in enumerate(range(min_col, max_col + 1)):
            celula = ws.cell(row=max_row, column=col)
            if celula.is_date or "yy" in str(celula.number_format).lower():
                colunas_data[j] = celula.number_format
        
        return {"tabela": tabela, "min_col": min_col, "min_row": min_row, 
                "max_col": max_col, "max_row": max_row, "headers": headers, "colunas_data": colunas_data}

    def escrever_dados_excel(ws, df, min_col, min_row, max_col, colunas_data):
        for i, row in enumerate(df.values, start=min_row):
            for j, value in enumerate(row):
                col_excel = min_col + j
                if j in colunas_data and isinstance(value, Timestamp):
                    value = None if pd.isna(value) else value.to_pydatetime()
                celula = ws.cell(row=i, column=col_excel, value=value)
                if j in colunas_data:
                    celula.number_format = colunas_data[j]

    def enviar_email(destinatario, cc):
        saudacao = "Bom dia, pessoal." if Timestamp.now().hour < 12 else "Boa tarde, pessoal."
        try:
            outlook = win32.Dispatch('outlook.application')
            mail = outlook.CreateItem(0)
        except Exception as e:
            logger.exception(f"Erro ao criar instância do Outlook: {e}")
            raise

        mail.Display()
        assinatura = mail.HtmlBody

        mail.to = destinatario
        mail.CC = cc
        mail.Subject = "Relatório SAC controle - Agências ALMAP BBDO, AFRICA, DPZ e MEDIABRANDS"

        mail.HtmlBody = Rf"""<div style="font-family: tahoma; font-size: 11pt">
            <p>{saudacao}</p>
            <p>A planilha com os questionamentos SAC (ALMAP BBDO, AFRICA, DPZ e MEDIABRANDS), está atualizada com dados até {Timestamp.now().strftime('%d/%m/%Y')}</p>
            <p>Arquivo: <a href="file:///\\Brmdfs05sp-fs.grupoibope.corp\groups\Monitor\ColetaTV\Tratamento Coleta TV\2026\AUXILIARES\Relatorio sac controle 2026 - agencias Almap e Africa.xlsx">G:\Monitor\ColetaTV\Tratamento Coleta TV‚6\AUXILIARES\Relatorio sac controle 2026 - agencias Almap e Africa.xlsx</a></p>
            <p style="font-family: tahoma; font-size: 9pt; color: #555;"><i>E-mail enviado automaticamente.</i></p>
            {assinatura}
        </div>
        """
        mail.Send()
        logger.info(f"E-mail enviado\n"
                    f"Destinatário: {destinatario}\n"
                    f"CC: {cc}")

    logger.info("Processo iniciado")

    # verifica se o arquivo está aberto
    arquivo = Path(planilha_final_path)

    lock_file = arquivo.parent / f"~${arquivo.name}"

    if lock_file.exists():
        logger.info("Planilha está aberta por alguém, tente executar mais tarde.")
        return
    else:
        logger.info("Planilha não está aberta por ninguém, processo seguirá.")

    # backup
    fazer_backup(planilha_final_path, planilha_baixada_path, logger)
    
    # Ler e processar dados
    try:
        df = pd.read_csv(planilha_baixada_path, encoding='latin1', sep=';')
        logger.info("Arquivo CSV lido com sucesso")
    except Exception as e:
        logger.exception(f"Erro ao ler o arquivo CSV: {e}")
        raise

    df.columns = df.columns.str.strip()

    try:
        wb = load_workbook(planilha_final_path)
        logger.info("Arquivo Excel lido com sucesso")
    except Exception as e:
        logger.exception(f"Erro ao ler o arquivo Excel: {e}")
        raise

    ws = wb.active
    info = obter_estrutura_tabela(ws)
    linhas_final_antes = len(wb.worksheets[0]['A'])

    headers = info["headers"]
    colunas_data = info["colunas_data"]

    meses_pt = {'jan':'01','fev':'02','mar':'03','abr':'04','mai':'05','jun':'06',
                'jul':'07','ago':'08','set':'09','out':'10','nov':'11','dez':'12'}

    # Garantir ordem e converter datas
    for col_name in df.columns:
        # Verificação para identificar colunas de data pelos nomes conhecidos (testar se é necessário)
        if "data" in col_name.lower() or "mês" in col_name.lower() or "veiculação" in col_name.lower():
            if col_name == "Mês Veiculação":
                for sigla, num in meses_pt.items():
                    df[col_name] = df[col_name].astype(str).str.replace(sigla, f"01/{num}", case=False, regex=False)
            
            df[col_name] = pd.to_datetime(df[col_name], errors='coerce', dayfirst=True, format='mixed')

    # Inserir novos dados
    escrever_dados_excel(ws, df, info["min_col"], info["max_row"] + 1, info["max_col"], colunas_data)
    nova_linha = info["max_row"] + len(df)
    info["tabela"].ref = f"{ws.cell(row=info['min_row'], column=info['min_col']).coordinate}:{ws.cell(row=nova_linha, column=info['max_col']).coordinate}"
    wb.save(planilha_final_path)
    logger.info(f"Linhas inseridas: {len(df)}")

    # Remover duplicados e reescrever
    try:
        df = pd.read_excel(planilha_final_path)
        logger.info("Removendo registros duplicados")
    except Exception as e:
        logger.exception(f"Erro ao ler o arquivo Excel: {e}")
        raise

    linhas_antes = len(df)
    df = df.drop_duplicates(subset=[c for c in df.columns if c != "Data Devolução"], keep='first')
    linhas_deletadas = linhas_antes - len(df)
    logger.info(f"Linhas duplicadas removidas: {linhas_deletadas}")
    df = df[headers]

    info = obter_estrutura_tabela(ws)
    ws.delete_rows(info["min_row"] + 1, ws.max_row)
    escrever_dados_excel(ws, df, info["min_col"], info["min_row"] + 1, info["max_col"], colunas_data)

    nova_linha = info["min_row"] + len(df)
    info["tabela"].ref = f"{get_column_letter(info['min_col'])}{info['min_row']}:{get_column_letter(info['max_col'])}{nova_linha}"

    linhas_final_depois = len(wb.worksheets[0]['A'])
    logger.info(f"Linhas novas: {linhas_final_depois - linhas_final_antes}")

    for tentativa in range(5):
        try:
            wb.save(planilha_final_path)
            logger.info("Planilha salva")
            break

        except Exception as e:
            logger.exception(f"Erro ao salvar o arquivo Excel: {e}")
            logger.info(f"Tentativa {tentativa + 1}/5 - arquivo bloqueado")
            time.sleep(10)

    else:    
        raise Exception("Não foi possível salvar a planilha após 5 tentativas.")
    
    enviar_email(destinatario, cc)
    logger.info(f"Processo finalizado. {time.perf_counter()-inicio:.2f}s")
